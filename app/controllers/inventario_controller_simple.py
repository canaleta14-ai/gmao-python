from app.models.inventario import Inventario
from app.models.movimiento_inventario import MovimientoInventario
from app.extensions import db
from flask import request
from datetime import datetime, timezone, timedelta
from sqlalchemy import func, text
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logger = logging.getLogger(__name__)


def obtener_estadisticas_inventario():
    """Obtener estadísticas generales del inventario"""
    stats = {}
    try:
        # Ruta principal (ORM)
        stats["total_articulos"] = Inventario.query.filter_by(activo=True).count()
        stats["articulos_bajo_minimo"] = Inventario.query.filter(
            Inventario.activo == True,
            Inventario.stock_actual <= Inventario.stock_minimo,
        ).count()
        valor_total = (
            db.session.query(
                func.sum(Inventario.stock_actual * Inventario.precio_promedio)
            )
            .filter(Inventario.activo == True)
            .scalar()
            or 0
        )
        stats["valor_total_stock"] = float(valor_total) if valor_total else 0
        stats["articulos_criticos"] = Inventario.query.filter_by(
            activo=True, critico=True
        ).count()
        return stats
    except Exception:
        # Evitar contaminación de la sesión del ORM en estado abortado
        try:
            db.session.rollback()
        except Exception:
            pass
        try:
            db.session.remove()
        except Exception:
            pass
        # Fallback: contar artículos totales y devolver el resto como 0
        total = 0
        try:
            backend = (
                db.engine.url.get_backend_name()
                if getattr(db, "engine", None)
                else None
            )
            table_name = "inventario"
            if backend == "postgresql":
                table_name = "public.inventario"
            from sqlalchemy import text as _text

            # Limpiar estado y ejecutar en AUTOCOMMIT para evitar transacción abortada
            with db.engine.connect() as base_conn:
                try:
                    base_conn.exec_driver_sql("ROLLBACK")
                except Exception:
                    pass
                with base_conn.execution_options(isolation_level="AUTOCOMMIT") as conn:
                    res = conn.execute(
                        _text(f"SELECT COUNT(*) AS total FROM {table_name}")
                    ).first()
                total = int(res[0]) if res else 0
        except Exception:
            total = 0
        return {
            "total_articulos": total,
            "articulos_bajo_minimo": 0,
            "valor_total_stock": 0,
            "articulos_criticos": 0,
        }


def listar_articulos_avanzado(filtros=None, page=1, per_page=10):
    """Listar artículos con filtros avanzados"""
    try:
        query = Inventario.query.filter_by(activo=True)

        if filtros:
            if "codigo" in filtros and filtros["codigo"]:
                query = query.filter(Inventario.codigo.ilike(f"%{filtros['codigo']}%"))
            if "descripcion" in filtros and filtros["descripcion"]:
                query = query.filter(
                    Inventario.descripcion.ilike(f"%{filtros['descripcion']}%")
                )
            if "categoria" in filtros and filtros["categoria"]:
                query = query.filter(Inventario.categoria == filtros["categoria"])
            if "bajo_minimo" in filtros and filtros["bajo_minimo"]:
                query = query.filter(Inventario.stock_actual <= Inventario.stock_minimo)
            if "critico" in filtros and filtros["critico"] == "true":
                query = query.filter(Inventario.critico == True)
            if "ubicacion" in filtros and filtros["ubicacion"]:
                query = query.filter(
                    Inventario.ubicacion.ilike(f"%{filtros['ubicacion']}%")
                )

            # Filtro de búsqueda general (para autocompletado)
            if "busqueda_general" in filtros and filtros["busqueda_general"]:
                busqueda = filtros["busqueda_general"]
                query = query.filter(
                    db.or_(
                        Inventario.codigo.ilike(f"%{busqueda}%"),
                        Inventario.descripcion.ilike(f"%{busqueda}%"),
                        Inventario.categoria.ilike(f"%{busqueda}%"),
                        Inventario.ubicacion.ilike(f"%{busqueda}%"),
                    )
                )

        paginacion = query.order_by(Inventario.codigo).paginate(
            page=page, per_page=per_page, error_out=False
        )
        return paginacion.items, paginacion.total
    except Exception:
        # Fallback robusto con SQL directo: usar columnas seguras
        try:
            backend = (
                db.engine.url.get_backend_name()
                if getattr(db, "engine", None)
                else None
            )
            table_name = "inventario"
            if backend == "postgresql":
                table_name = "public.inventario"

            base_sql = f"SELECT id, codigo, descripcion FROM {table_name}"
            where_clauses = []
            params = {}

            if filtros:
                if filtros.get("codigo"):
                    where_clauses.append("codigo ILIKE :codigo")
                    params["codigo"] = f"%{filtros['codigo']}%"
                if filtros.get("descripcion"):
                    where_clauses.append("descripcion ILIKE :descripcion")
                    params["descripcion"] = f"%{filtros['descripcion']}%"
                if filtros.get("busqueda_general"):
                    where_clauses.append("(codigo ILIKE :q OR descripcion ILIKE :q)")
                    params["q"] = f"%{filtros['busqueda_general']}%"

            if where_clauses:
                base_sql += " WHERE " + " AND ".join(where_clauses)

            base_sql += " ORDER BY codigo LIMIT :limit OFFSET :offset"
            params["limit"] = per_page
            params["offset"] = (page - 1) * per_page

            count_sql = f"SELECT COUNT(*) AS total FROM {table_name}"
            if where_clauses:
                count_sql += " WHERE " + " AND ".join(where_clauses)

            # Usar AUTOCOMMIT tras un ROLLBACK explícito para evitar estados abortados
            with db.engine.connect() as base_conn:
                try:
                    base_conn.exec_driver_sql("ROLLBACK")
                except Exception:
                    pass
                with base_conn.execution_options(isolation_level="AUTOCOMMIT") as conn:
                    rows = conn.execute(text(base_sql), params).mappings().all()
                    total_row = conn.execute(text(count_sql), params).first()
            total = total_row[0] if total_row else 0

            articulos = [dict(r) for r in rows]
            return articulos, total
        except Exception:
            # Si también falla el fallback (tabla inexistente, etc.), devolver vacío
            return [], 0


def crear_articulo_simple(data):
    """Crear un nuevo artículo de inventario simple"""
    # Validaciones básicas
    if not data.get("codigo"):
        raise ValueError("El código es obligatorio")
    if not data.get("descripcion"):
        raise ValueError("La descripción es obligatoria")

    # Validar código único
    if Inventario.query.filter_by(codigo=data["codigo"]).first():
        raise ValueError("Ya existe un artículo con este código")

    # Validar y convertir valores numéricos
    try:
        stock_actual = float(data.get("stock_actual", 0))
        stock_minimo = float(data.get("stock_minimo", 0))
        stock_maximo = float(data.get("stock_maximo", 100))
        precio_unitario = float(data.get("precio_unitario", 0))
        precio_promedio = float(data.get("precio_promedio", 0))
    except (ValueError, TypeError):
        raise ValueError("Los valores numéricos deben ser válidos")

    # Validar que los valores no sean negativos
    if stock_actual < 0:
        raise ValueError("El stock actual no puede ser negativo")
    if stock_minimo < 0:
        raise ValueError("El stock mínimo no puede ser negativo")
    if stock_maximo < 0:
        raise ValueError("El stock máximo no puede ser negativo")
    if precio_unitario < 0:
        raise ValueError("El precio unitario no puede ser negativo")
    if precio_promedio < 0:
        raise ValueError("El precio promedio no puede ser negativo")

    # Validar lógica de stocks
    if stock_maximo > 0 and stock_minimo > stock_maximo:
        raise ValueError("El stock mínimo no puede ser mayor que el stock máximo")

    articulo = Inventario(
        codigo=data["codigo"],
        descripcion=data["descripcion"],
        categoria=data.get("categoria", ""),
        stock_actual=stock_actual,
        stock_minimo=stock_minimo,
        stock_maximo=stock_maximo,
        ubicacion=data.get("ubicacion", ""),
        precio_unitario=precio_unitario,
        precio_promedio=precio_promedio,
        unidad_medida=data.get("unidad_medida", "UNI"),
        proveedor_principal=data.get("proveedor_principal", ""),
        grupo_contable=data.get("grupo_contable", ""),
        cuenta_contable_compra=data.get("cuenta_contable_compra", "622000000"),
        critico=data.get("critico", False),
        activo=data.get("activo", True),
        observaciones=data.get("observaciones", ""),
    )

    db.session.add(articulo)
    db.session.commit()
    return articulo


def exportar_inventario_csv():
    """Genera un archivo Excel con todos los artículos del inventario"""
    articulos = Inventario.query.filter_by(activo=True).all()

    # Crear un nuevo workbook de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = [
        "Código",
        "Descripción",
        "Categoría",
        "Stock Actual",
        "Stock Mínimo",
        "Stock Máximo",
        "Ubicación",
        "Precio Unitario",
        "Precio Promedio",
        "Valor Stock",
        "Unidad Medida",
        "Proveedor Principal",
        "Cuenta Contable",
        "Grupo Contable",
        "Crítico",
        "Fecha Creación",
    ]

    # Aplicar estilos al encabezado
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ajustar ancho de columnas
    column_widths = [15, 40, 15, 12, 12, 12, 20, 15, 15, 15, 12, 25, 15, 15, 8, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Escribir datos
    for row_num, articulo in enumerate(articulos, 2):
        ws.cell(row=row_num, column=1, value=articulo.codigo)
        ws.cell(row=row_num, column=2, value=articulo.descripcion)
        ws.cell(row=row_num, column=3, value=articulo.categoria or "")
        ws.cell(row=row_num, column=4, value=articulo.stock_actual)
        ws.cell(row=row_num, column=5, value=articulo.stock_minimo)
        ws.cell(row=row_num, column=6, value=articulo.stock_maximo or "")
        ws.cell(row=row_num, column=7, value=articulo.ubicacion or "")
        ws.cell(
            row=row_num,
            column=8,
            value=float(articulo.precio_unitario) if articulo.precio_unitario else "",
        )
        ws.cell(
            row=row_num,
            column=9,
            value=float(articulo.precio_promedio) if articulo.precio_promedio else "",
        )
        ws.cell(
            row=row_num,
            column=10,
            value=(
                float(articulo.stock_actual * articulo.precio_promedio)
                if articulo.precio_promedio
                else 0
            ),
        )
        ws.cell(row=row_num, column=11, value=articulo.unidad_medida)
        ws.cell(row=row_num, column=12, value=articulo.proveedor_principal or "")
        ws.cell(row=row_num, column=13, value=articulo.cuenta_contable_compra)
        ws.cell(row=row_num, column=14, value=articulo.grupo_contable or "")
        ws.cell(row=row_num, column=15, value="Sí" if articulo.critico else "No")
        ws.cell(
            row=row_num,
            column=16,
            value=(
                articulo.fecha_creacion.strftime("%d/%m/%Y")
                if articulo.fecha_creacion
                else ""
            ),
        )

    # Guardar el workbook en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# Funciones para movimientos de inventario
from app.models.movimiento_inventario import MovimientoInventario
from app.models.usuario import Usuario


def registrar_movimiento_inventario(data):
    """Registrar un nuevo movimiento de inventario"""
    from app.services.servicio_fifo import ServicioFIFO

    try:
        # Validar datos requeridos
        required_fields = ["inventario_id", "tipo", "cantidad"]
        for field in required_fields:
            if field not in data:
                raise ValueError(f"Campo requerido faltante: {field}")

        # Obtener el artículo
        articulo = Inventario.query.get(data["inventario_id"])
        if not articulo:
            raise ValueError("Artículo no encontrado")

        # Crear el movimiento
        movimiento = MovimientoInventario(
            inventario_id=data["inventario_id"],
            tipo=data["tipo"],
            subtipo=data.get("subtipo"),
            cantidad=data["cantidad"],
            precio_unitario=data.get("precio_unitario"),
            valor_total=data.get("valor_total"),
            cuenta_contable=data.get(
                "cuenta_contable", articulo.cuenta_contable_compra
            ),
            centro_costo=data.get("centro_costo"),
            documento_referencia=data.get("documento_referencia"),
            observaciones=data.get("observaciones"),
            usuario_id=data.get("usuario_id", "sistema"),
        )

        # Actualizar stock del artículo según el tipo de movimiento
        stock_anterior = float(articulo.stock_actual)

        if movimiento.tipo == "entrada":
            articulo.stock_actual += movimiento.cantidad

            # 🆕 CREACIÓN AUTOMÁTICA DE LOTE FIFO para entradas
            try:
                # Generar código de lote automático
                fecha_hoy = datetime.now()
                codigo_lote = f"{articulo.codigo}-{fecha_hoy.strftime('%Y%m%d')}-{movimiento.id or 'AUTO'}"

                # Calcular fecha de vencimiento estimada (si no se proporciona)
                fecha_vencimiento = data.get("fecha_vencimiento")
                if (
                    not fecha_vencimiento
                    and hasattr(articulo, "categoria")
                    and articulo.categoria
                ):
                    # Asignar vencimiento por defecto según categoría
                    dias_vencimiento = {
                        "medicamentos": 365,
                        "quimicos": 180,
                        "repuestos": 1095,  # 3 años
                        "consumibles": 365,
                        "herramientas": 1825,  # 5 años
                    }.get(
                        str(articulo.categoria).lower(), 730
                    )  # 2 años por defecto

                    fecha_vencimiento = fecha_hoy + timedelta(days=dias_vencimiento)

                # Crear lote FIFO automáticamente
                lote_fifo = ServicioFIFO.crear_lote_entrada(
                    inventario_id=articulo.id,
                    cantidad=abs(movimiento.cantidad),
                    precio_unitario=movimiento.precio_unitario or 0,
                    codigo_lote=codigo_lote,
                    fecha_vencimiento=fecha_vencimiento,
                    documento_origen=movimiento.documento_referencia,
                    proveedor_id=getattr(movimiento, "proveedor_id", None),
                    usuario_id=movimiento.usuario_id or "sistema",
                    observaciones=f"Lote creado automáticamente por entrada de inventario. {movimiento.observaciones or ''}",
                )

                logger.info(
                    f"✅ Lote FIFO creado automáticamente: {lote_fifo.id} para artículo {articulo.codigo}"
                )

            except Exception as e:
                logger.error(f"❌ Error al crear lote FIFO automático: {str(e)}")
                # No fallar la entrada por error en lote, solo registrar

            # Actualizar precio promedio ponderado si hay precio unitario
            if movimiento.precio_unitario and movimiento.precio_unitario > 0:
                valor_stock_anterior = stock_anterior * float(
                    articulo.precio_promedio or 0
                )
                valor_entrada = float(movimiento.cantidad) * float(
                    movimiento.precio_unitario
                )
                valor_total_nuevo = valor_stock_anterior + valor_entrada

                if articulo.stock_actual > 0:
                    articulo.precio_promedio = valor_total_nuevo / float(
                        articulo.stock_actual
                    )
                else:
                    articulo.precio_promedio = float(movimiento.precio_unitario)

        elif movimiento.tipo == "salida":
            if articulo.stock_actual < movimiento.cantidad:
                raise ValueError("Stock insuficiente para la salida")

            # 🆕 CONSUMO AUTOMÁTICO FIFO para salidas
            try:
                # Usar ServicioFIFO para consumir automáticamente
                consumos, cantidad_faltante = ServicioFIFO.consumir_fifo(
                    inventario_id=articulo.id,
                    cantidad_total=abs(movimiento.cantidad),
                    orden_trabajo_id=getattr(movimiento, "orden_trabajo_id", None),
                    documento_referencia=movimiento.documento_referencia,
                    usuario_id=movimiento.usuario_id or "sistema",
                    observaciones=f"Consumo automático por salida de inventario. {movimiento.observaciones or ''}",
                )

                if cantidad_faltante > 0:
                    logger.warning(
                        f"⚠️ Cantidad faltante en FIFO: {cantidad_faltante} para artículo {articulo.codigo}"
                    )

                logger.info(
                    f"✅ Consumo FIFO automático: {len(consumos)} lotes afectados para artículo {articulo.codigo}"
                )

            except Exception as e:
                logger.error(f"❌ Error en consumo FIFO automático: {str(e)}")
                # Continuar con el método tradicional

            articulo.stock_actual -= movimiento.cantidad
            # Para salidas, mantener el precio promedio (no cambiar)

        elif movimiento.tipo == "ajuste":
            # Para ajustes, la cantidad puede ser positiva o negativa
            articulo.stock_actual += movimiento.cantidad

            # Si es ajuste positivo con precio, actualizar precio promedio
            if (
                movimiento.cantidad > 0
                and movimiento.precio_unitario
                and movimiento.precio_unitario > 0
            ):
                valor_stock_anterior = stock_anterior * float(
                    articulo.precio_promedio or 0
                )
                valor_ajuste = float(movimiento.cantidad) * float(
                    movimiento.precio_unitario
                )
                valor_total_nuevo = valor_stock_anterior + valor_ajuste

                if articulo.stock_actual > 0:
                    articulo.precio_promedio = valor_total_nuevo / float(
                        articulo.stock_actual
                    )

        elif movimiento.tipo == "regularizacion":
            # Regularización suma al stock (corrige inventario)
            articulo.stock_actual += movimiento.cantidad

            # Si hay precio, actualizar precio promedio
            if movimiento.precio_unitario and movimiento.precio_unitario > 0:
                valor_stock_anterior = stock_anterior * float(
                    articulo.precio_promedio or 0
                )
                valor_regularizacion = float(movimiento.cantidad) * float(
                    movimiento.precio_unitario
                )
                valor_total_nuevo = valor_stock_anterior + valor_regularizacion

                if articulo.stock_actual > 0:
                    articulo.precio_promedio = valor_total_nuevo / float(
                        articulo.stock_actual
                    )

        # Calcular valor total si no se proporciona
        if not movimiento.valor_total and movimiento.precio_unitario:
            movimiento.valor_total = movimiento.precio_unitario * abs(
                movimiento.cantidad
            )

        db.session.add(movimiento)
        db.session.commit()

        return movimiento

    except Exception as e:
        db.session.rollback()
        raise e


def obtener_movimientos_articulo(articulo_id, page=1, per_page=10):
    """Obtener historial de movimientos de un artículo"""
    try:
        query = MovimientoInventario.query.filter_by(inventario_id=articulo_id)
        query = query.order_by(MovimientoInventario.fecha.desc())

        # Paginación
        movimientos = query.paginate(page=page, per_page=per_page, error_out=False)

        return {
            "movimientos": [
                {
                    "id": m.id,
                    "fecha": m.fecha.strftime("%d/%m/%Y %H:%M") if m.fecha else "",
                    "tipo": m.tipo,
                    "subtipo": m.subtipo,
                    "cantidad": m.cantidad,
                    "precio_unitario": m.precio_unitario,
                    "valor_total": m.valor_total,
                    "cuenta_contable": m.cuenta_contable,
                    "documento_referencia": m.documento_referencia,
                    "observaciones": m.observaciones,
                    "usuario_id": m.usuario_id,
                }
                for m in movimientos.items
            ],
            "page": movimientos.page,
            "per_page": movimientos.per_page,
            "total": movimientos.total,
            "pages": movimientos.pages,
        }

    except Exception as e:
        raise e


def obtener_movimientos_generales(filtros=None, page=1, per_page=10):
    """Obtener vista general de movimientos con filtros"""
    try:
        query = MovimientoInventario.query

        # Aplicar filtros
        if filtros:
            if "tipo" in filtros:
                query = query.filter_by(tipo=filtros["tipo"])
            if "fecha_desde" in filtros:
                query = query.filter(
                    MovimientoInventario.fecha >= filtros["fecha_desde"]
                )
            if "fecha_hasta" in filtros:
                query = query.filter(
                    MovimientoInventario.fecha <= filtros["fecha_hasta"]
                )
            if "usuario_id" in filtros:
                query = query.filter_by(usuario_id=filtros["usuario_id"])

        # Unir con información del artículo
        query = query.join(Inventario).add_columns(
            Inventario.codigo, Inventario.descripcion, Inventario.unidad_medida
        )

        query = query.order_by(MovimientoInventario.fecha.desc())

        # Paginación
        movimientos = query.paginate(page=page, per_page=per_page, error_out=False)

        return {
            "movimientos": [
                {
                    "id": m.MovimientoInventario.id,
                    "fecha": (
                        m.MovimientoInventario.fecha.strftime("%d/%m/%Y %H:%M")
                        if m.MovimientoInventario.fecha
                        else ""
                    ),
                    "tipo": m.MovimientoInventario.tipo,
                    "subtipo": m.MovimientoInventario.subtipo,
                    "cantidad": m.MovimientoInventario.cantidad,
                    "codigo_articulo": m.codigo,
                    "descripcion_articulo": m.descripcion,
                    "unidad_medida": m.unidad_medida,
                    "precio_unitario": m.MovimientoInventario.precio_unitario,
                    "valor_total": m.MovimientoInventario.valor_total,
                    "documento_referencia": m.MovimientoInventario.documento_referencia,
                    "usuario_id": m.MovimientoInventario.usuario_id,
                }
                for m in movimientos.items
            ],
            "page": movimientos.page,
            "per_page": movimientos.per_page,
            "total": movimientos.total,
            "pages": movimientos.pages,
        }

    except Exception as e:
        raise e


def editar_articulo_simple(articulo_id, data):
    """Editar un artículo existente"""
    try:
        articulo = Inventario.query.get(articulo_id)
        if not articulo:
            raise ValueError("Artículo no encontrado")

        # Campos que se pueden editar
        campos_editables = [
            "codigo",
            "descripcion",
            "categoria",
            "subcategoria",
            "ubicacion",
            "stock_minimo",
            "stock_maximo",
            "unidad_medida",
            "precio_unitario",
            "proveedor_principal",
            "cuenta_contable_compra",
            "grupo_contable",
            "critico",
            "activo",
            "observaciones",
        ]

        for campo in campos_editables:
            if campo in data:
                setattr(articulo, campo, data[campo])

        articulo.fecha_actualizacion = datetime.now(timezone.utc)

        db.session.commit()
        return articulo

    except Exception as e:
        db.session.rollback()
        raise e


def eliminar_articulo(articulo_id):
    """Eliminar un artículo del inventario"""
    try:
        articulo = Inventario.query.get(articulo_id)
        if not articulo:
            raise ValueError("Artículo no encontrado")

        db.session.delete(articulo)
        db.session.commit()
        return True

    except Exception as e:
        db.session.rollback()
        raise e
