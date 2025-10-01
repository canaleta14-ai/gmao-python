from app.models.proveedor import Proveedor
from app.extensions import db
from flask import jsonify
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def listar_proveedores(filtros=None):
    """Retorna todos los proveedores con soporte para filtros de búsqueda"""
    query = Proveedor.query

    # Aplicar filtros de búsqueda si se proporcionan
    if filtros:
        if "nombre" in filtros and filtros["nombre"]:
            query = query.filter(Proveedor.nombre.ilike(f"%{filtros['nombre']}%"))

        if "nif" in filtros and filtros["nif"]:
            query = query.filter(Proveedor.nif.ilike(f"%{filtros['nif']}%"))

        if "contacto" in filtros and filtros["contacto"]:
            query = query.filter(Proveedor.contacto.ilike(f"%{filtros['contacto']}%"))

        if "q" in filtros and filtros["q"]:
            # Búsqueda general en múltiples campos
            search_term = f"%{filtros['q']}%"
            query = query.filter(
                db.or_(
                    Proveedor.nombre.ilike(search_term),
                    Proveedor.nif.ilike(search_term),
                    Proveedor.contacto.ilike(search_term),
                    Proveedor.email.ilike(search_term),
                )
            )

    # Limitar resultados si se especifica
    if filtros and "limit" in filtros:
        try:
            limit = int(filtros["limit"])
            query = query.limit(limit)
        except (ValueError, TypeError):
            pass

    proveedores = query.all()
    return [
        {
            "id": p.id,
            "nombre": p.nombre,
            "nif": p.nif,
            "cuenta_contable": p.cuenta_contable,
            "direccion": p.direccion,
            "telefono": p.telefono,
            "email": p.email,
            "contacto": p.contacto,
            "activo": p.activo,
        }
        for p in proveedores
    ]


def listar_proveedores_paginado(page=1, per_page=10, q=None):
    """Listar proveedores con paginación y filtros"""
    query = Proveedor.query

    # Filtro de búsqueda general
    if q:
        search_term = f"%{q}%"
        query = query.filter(
            db.or_(
                Proveedor.nombre.ilike(search_term),
                Proveedor.nif.ilike(search_term),
                Proveedor.contacto.ilike(search_term),
                Proveedor.email.ilike(search_term),
                Proveedor.direccion.ilike(search_term),
                Proveedor.telefono.ilike(search_term),
            )
        )

    # Ordenamiento por nombre
    query = query.order_by(Proveedor.nombre.asc())

    # Paginación
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    proveedores_data = [
        {
            "id": p.id,
            "nombre": p.nombre,
            "nif": p.nif,
            "cuenta_contable": p.cuenta_contable,
            "direccion": p.direccion,
            "telefono": p.telefono,
            "email": p.email,
            "contacto": p.contacto,
            "activo": p.activo,
        }
        for p in pagination.items
    ]

    return {
        "items": proveedores_data,
        "page": pagination.page,
        "per_page": pagination.per_page,
        "total": pagination.total,
        "pages": pagination.pages,
        "has_next": pagination.has_next,
        "has_prev": pagination.has_prev,
    }


def crear_proveedor(data):
    """Crea un nuevo proveedor"""
    # Validaciones básicas
    if not data.get("nombre"):
        raise ValueError("El nombre es requerido")

    if not data.get("nif"):
        raise ValueError("El NIF es requerido")

    if not data.get("cuenta_contable"):
        raise ValueError("La cuenta contable es requerida")

    # Verificar que el NIF no esté duplicado
    proveedor_existente = Proveedor.query.filter_by(nif=data["nif"]).first()
    if proveedor_existente:
        raise ValueError("Ya existe un proveedor con este NIF")

    # Crear nuevo proveedor
    nuevo_proveedor = Proveedor(
        nombre=data["nombre"].strip(),
        nif=data["nif"].strip().upper(),
        cuenta_contable=data["cuenta_contable"].strip(),
        direccion=data.get("direccion", "").strip(),
        telefono=data.get("telefono", "").strip(),
        email=data.get("email", "").strip(),
        contacto=data.get("contacto", "").strip(),
        activo=True,
    )

    db.session.add(nuevo_proveedor)
    db.session.commit()

    return nuevo_proveedor


def actualizar_proveedor(id, data):
    """Actualiza un proveedor existente"""
    proveedor = Proveedor.query.get_or_404(id)

    # Validaciones básicas
    if not data.get("nombre"):
        raise ValueError("El nombre es requerido")

    if not data.get("nif"):
        raise ValueError("El NIF es requerido")

    if not data.get("cuenta_contable"):
        raise ValueError("La cuenta contable es requerida")

    # Verificar que el NIF no esté duplicado (excepto para el mismo proveedor)
    proveedor_existente = Proveedor.query.filter(
        Proveedor.nif == data["nif"].strip().upper(), Proveedor.id != id
    ).first()
    if proveedor_existente:
        raise ValueError("Ya existe otro proveedor con este NIF")

    # Actualizar campos
    proveedor.nombre = data["nombre"].strip()
    proveedor.nif = data["nif"].strip().upper()
    proveedor.cuenta_contable = data["cuenta_contable"].strip()
    proveedor.direccion = data.get("direccion", "").strip()
    proveedor.telefono = data.get("telefono", "").strip()
    proveedor.email = data.get("email", "").strip()
    proveedor.contacto = data.get("contacto", "").strip()

    db.session.commit()

    return proveedor


def eliminar_proveedor(id):
    """Marca un proveedor como inactivo (soft delete)"""
    proveedor = Proveedor.query.get_or_404(id)
    proveedor.activo = False
    db.session.commit()
    return True


def obtener_proveedor(id):
    """Obtiene un proveedor específico"""
    proveedor = Proveedor.query.get_or_404(id)
    return {
        "id": proveedor.id,
        "nombre": proveedor.nombre,
        "nif": proveedor.nif,
        "cuenta_contable": proveedor.cuenta_contable,
        "direccion": proveedor.direccion,
        "telefono": proveedor.telefono,
        "email": proveedor.email,
        "contacto": proveedor.contacto,
        "activo": proveedor.activo,
    }


def exportar_proveedores_csv():
    """Genera un archivo Excel con todos los proveedores activos"""
    proveedores = Proveedor.query.filter_by(activo=True).all()

    # Crear un nuevo workbook de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    # Estilos para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = [
        "ID",
        "Nombre",
        "NIF",
        "Cuenta Contable",
        "Dirección",
        "Teléfono",
        "Email",
        "Contacto",
    ]

    # Aplicar estilos al encabezado
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ajustar ancho de columnas
    column_widths = [8, 30, 12, 15, 40, 15, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Escribir datos
    for row_num, proveedor in enumerate(proveedores, 2):
        ws.cell(row=row_num, column=1, value=proveedor.id)
        ws.cell(row=row_num, column=2, value=proveedor.nombre)
        ws.cell(row=row_num, column=3, value=proveedor.nif)
        ws.cell(row=row_num, column=4, value=proveedor.cuenta_contable)
        ws.cell(row=row_num, column=5, value=proveedor.direccion or "")
        ws.cell(row=row_num, column=6, value=proveedor.telefono or "")
        ws.cell(row=row_num, column=7, value=proveedor.email or "")
        ws.cell(row=row_num, column=8, value=proveedor.contacto or "")

    # Guardar el workbook en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def validar_nif(nif):
    """Valida que un NIF no esté duplicado"""
    proveedor_existente = Proveedor.query.filter_by(nif=nif.strip().upper()).first()
    if proveedor_existente:
        return {"valido": False, "mensaje": "El NIF ya existe"}
    return {"valido": True, "mensaje": "NIF disponible"}


def obtener_estadisticas_proveedores():
    """Obtiene estadísticas generales de los proveedores"""
    from sqlalchemy import func

    # Contar total de proveedores
    total_proveedores = Proveedor.query.count()

    # Contar por estado (activo/inactivo)
    estados = (
        db.session.query(Proveedor.activo, func.count(Proveedor.id).label("count"))
        .group_by(Proveedor.activo)
        .all()
    )

    # Inicializar contadores
    proveedores_activos = 0
    proveedores_inactivos = 0

    # Procesar los resultados
    for activo, count in estados:
        if activo:
            proveedores_activos = count
        else:
            proveedores_inactivos = count

    # Para "pendientes", asumiremos que son proveedores inactivos que podrían estar en evaluación
    # Por ahora, pondremos 0 o podríamos agregar un campo adicional al modelo si es necesario
    proveedores_pendientes = 0

    return {
        "total_proveedores": total_proveedores,
        "proveedores_activos": proveedores_activos,
        "proveedores_pendientes": proveedores_pendientes,
        "proveedores_inactivos": proveedores_inactivos,
    }


def toggle_proveedor(id):
    """Activa o desactiva un proveedor"""
    proveedor = Proveedor.query.get_or_404(id)
    proveedor.activo = not proveedor.activo
    db.session.commit()
    return proveedor
