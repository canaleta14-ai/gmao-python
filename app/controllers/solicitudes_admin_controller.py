from flask import (
    Blueprint,
    render_template,
    request,
    flash,
    redirect,
    url_for,
    jsonify,
    Response,
)
from flask_login import login_required, current_user
from functools import wraps
from app.extensions import db
from app.models.solicitud_servicio import SolicitudServicio
from app.models.usuario import Usuario
from app.models.archivo_adjunto import ArchivoAdjunto
from app.utils.email_utils import enviar_email
from datetime import datetime, timezone
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def login_required_ajax(f):
    """Decorador personalizado para endpoints AJAX que devuelve JSON en lugar de redirigir"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return (
                jsonify(
                    {
                        "error": "Sesión expirada. Por favor, recarga la página e inicia sesión nuevamente."
                    }
                ),
                401,
            )
        return f(*args, **kwargs)

    return decorated_function


# Crear blueprint para gestión administrativa de solicitudes
solicitudes_admin_bp = Blueprint(
    "solicitudes_admin", __name__, url_prefix="/admin/solicitudes"
)


@solicitudes_admin_bp.route("/", methods=["GET"])
@login_required
def listar_solicitudes():
    """Lista todas las solicitudes para administración"""
    # Permitir acceso a Administradores, Técnicos y Supervisores
    if current_user.rol not in ["Administrador", "Técnico", "Supervisor"]:
        flash("No tiene permisos para acceder a esta sección.", "error")
        return redirect(url_for("web.index"))

    # Obtener filtros
    estado = request.args.get("estado", "")
    prioridad = request.args.get("prioridad", "")
    tipo_servicio = request.args.get("tipo_servicio", "")
    busqueda = request.args.get("busqueda", "").strip()
    page = int(request.args.get("page", 1))
    per_page = 25

    # Construir query
    query = SolicitudServicio.query

    if estado:
        query = query.filter_by(estado=estado)

    if prioridad:
        query = query.filter_by(prioridad=prioridad)

    if tipo_servicio:
        query = query.filter_by(tipo_servicio=tipo_servicio)

    if busqueda:
        # Buscar en número de solicitud, nombre, email o teléfono
        query = query.filter(
            db.or_(
                SolicitudServicio.numero_solicitud.ilike(f"%{busqueda}%"),
                SolicitudServicio.nombre_solicitante.ilike(f"%{busqueda}%"),
                SolicitudServicio.email_solicitante.ilike(f"%{busqueda}%"),
                SolicitudServicio.telefono_solicitante.ilike(f"%{busqueda}%"),
            )
        )

    # Ordenar por fecha de creación (más recientes primero)
    query = query.order_by(SolicitudServicio.fecha_creacion.desc())

    # Paginación
    solicitudes = query.paginate(page=page, per_page=per_page, error_out=False)

    # Calcular estadísticas
    estadisticas = {
        "total_pendientes": SolicitudServicio.query.filter_by(
            estado="pendiente"
        ).count(),
        "total_revision": SolicitudServicio.query.filter_by(
            estado="en_revision"
        ).count(),
        "total_progreso": SolicitudServicio.query.filter_by(
            estado="en_progreso"
        ).count(),
        "total_completadas": SolicitudServicio.query.filter_by(
            estado="completada"
        ).count(),
    }

    return render_template(
        "admin/solicitudes/listar.html",
        solicitudes=solicitudes,
        estadisticas=estadisticas,
        estado=estado,
        prioridad=prioridad,
        tipo_servicio=tipo_servicio,
        busqueda=busqueda,
        total_paginas=solicitudes.pages,
    )


@solicitudes_admin_bp.route("/<int:id>", methods=["GET"])
@login_required
def ver_solicitud(id):
    """Ver detalles de una solicitud específica"""
    if current_user.rol not in ["Administrador", "Técnico", "Supervisor"]:
        flash("No tiene permisos para acceder a esta sección.", "error")
        return redirect(url_for("web.index"))

    solicitud = SolicitudServicio.query.get_or_404(id)
    archivos_adjuntos = ArchivoAdjunto.query.filter_by(solicitud_servicio_id=solicitud.id).all()

    return render_template("admin/solicitudes/ver.html", solicitud=solicitud, archivos_adjuntos=archivos_adjuntos)


@solicitudes_admin_bp.route("/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_solicitud(id):
    """Editar una solicitud específica"""
    if current_user.rol not in ["Administrador", "Técnico", "Supervisor"]:
        flash("No tiene permisos para acceder a esta sección.", "error")
        return redirect(url_for("web.index"))

    solicitud = SolicitudServicio.query.get_or_404(id)

    if request.method == "POST":
        # Procesar el formulario de edición
        try:
            # Actualizar campos editables
            solicitud.titulo = request.form.get("titulo", solicitud.titulo)
            solicitud.descripcion = request.form.get(
                "descripcion", solicitud.descripcion
            )
            solicitud.tipo_servicio = request.form.get(
                "tipo_servicio", solicitud.tipo_servicio
            )
            solicitud.prioridad = request.form.get("prioridad", solicitud.prioridad)
            solicitud.fecha_actualizacion = datetime.utcnow()

            db.session.commit()

            flash("Solicitud actualizada correctamente.", "success")
            return redirect(url_for("solicitudes_admin.ver_solicitud", id=solicitud.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Error al actualizar la solicitud: {str(e)}", "error")

    return render_template("admin/solicitudes/editar.html", solicitud=solicitud)


@solicitudes_admin_bp.route("/<int:id>/estado", methods=["POST"])
@login_required
def cambiar_estado(id):
    """Cambiar el estado de una solicitud"""
    if current_user.rol != "Administrador":
        return jsonify({"error": "No autorizado"}), 403

    solicitud = SolicitudServicio.query.get_or_404(id)

    # Manejar tanto datos de formulario como JSON
    if request.is_json:
        data = request.get_json()
        nuevo_estado = data.get("estado")
        observaciones = data.get("comentario", "")  # Los templates usan "comentario"
    else:
        nuevo_estado = request.form.get("estado")
        observaciones = request.form.get("observaciones", "")

    if nuevo_estado not in [
        "pendiente",
        "en_revision",
        "aprobada",
        "rechazada",
        "en_progreso",
        "completada",
        "cancelada",
    ]:
        return jsonify({"error": "Estado no válido"}), 400

    estado_anterior = solicitud.estado
    solicitud.estado = nuevo_estado
    solicitud.fecha_actualizacion = datetime.utcnow()

    if observaciones:
        solicitud.observaciones_internas = observaciones

    try:
        db.session.commit()

        # Enviar notificación por email al solicitante
        enviar_notificacion_estado(solicitud, estado_anterior)

        return jsonify(
            {
                "success": True,
                "mensaje": f"Estado cambiado a {solicitud.estado_display}",
                "estado": solicitud.estado,
                "estado_display": solicitud.estado_display,
            }
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@solicitudes_admin_bp.route("/<int:id>/asignar", methods=["POST"])
@login_required
def asignar_solicitud(id):
    """Asignar una solicitud a un técnico"""
    if current_user.rol != "Administrador":
        return jsonify({"error": "No autorizado"}), 403

    solicitud = SolicitudServicio.query.get_or_404(id)

    tecnico_id = request.form.get("tecnico_id")
    if tecnico_id:
        tecnico = Usuario.query.get(int(tecnico_id))
        if tecnico and tecnico.rol in ["Administrador", "Técnico"]:
            solicitud.asignado_a_id = int(tecnico_id)
        else:
            return jsonify({"error": "Técnico no válido"}), 400
    else:
        solicitud.asignado_a_id = None

    solicitud.fecha_actualizacion = datetime.utcnow()

    try:
        db.session.commit()
        return jsonify(
            {
                "success": True,
                "mensaje": "Solicitud asignada correctamente",
                "asignado_a": (
                    solicitud.asignado_a.nombre if solicitud.asignado_a else None
                ),
            }
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@solicitudes_admin_bp.route("/api/estadisticas")
@login_required
def api_estadisticas():
    """API para obtener estadísticas de solicitudes"""
    if current_user.rol != "Administrador":
        return jsonify({"error": "No autorizado"}), 403

    # Estadísticas generales
    total = SolicitudServicio.query.count()
    pendientes = SolicitudServicio.query.filter_by(estado="pendiente").count()
    en_progreso = SolicitudServicio.query.filter_by(estado="en_progreso").count()
    completadas = SolicitudServicio.query.filter_by(estado="completada").count()

    # Por tipo de servicio
    tipos = (
        db.session.query(
            SolicitudServicio.tipo_servicio,
            db.func.count(SolicitudServicio.id).label("cantidad"),
        )
        .group_by(SolicitudServicio.tipo_servicio)
        .all()
    )

    # Por prioridad
    prioridades = (
        db.session.query(
            SolicitudServicio.prioridad,
            db.func.count(SolicitudServicio.id).label("cantidad"),
        )
        .group_by(SolicitudServicio.prioridad)
        .all()
    )

    return jsonify(
        {
            "total": total,
            "pendientes": pendientes,
            "en_progreso": en_progreso,
            "completadas": completadas,
            "tipos": [{"tipo": t[0], "cantidad": t[1]} for t in tipos],
            "prioridades": [{"prioridad": p[0], "cantidad": p[1]} for p in prioridades],
        }
    )


def enviar_notificacion_estado(solicitud, estado_anterior):
    """Enviar notificación por email cuando cambia el estado"""
    try:
        asunto = f"Actualización de Solicitud #{solicitud.numero_solicitud} - GMAO"

        contenido_html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <h2 style="color: #333;">Actualización de Estado de Solicitud</h2>

            <p>Estimado/a {solicitud.nombre_solicitante},</p>

            <p>Le informamos que el estado de su solicitud ha sido actualizado:</p>

            <div style="background-color: #f8f9fa; padding: 20px; border-radius: 5px; margin: 20px 0;">
                <h3>Detalles de la Solicitud</h3>
                <p><strong>Número de Solicitud:</strong> {solicitud.numero_solicitud}</p>
                <p><strong>Título:</strong> {solicitud.titulo}</p>
                <p><strong>Estado Anterior:</strong> {SolicitudServicio(estado=estado_anterior).estado_display}</p>
                <p><strong>Nuevo Estado:</strong> <span style="color: #0d6efd; font-weight: bold;">{solicitud.estado_display}</span></p>
                <p><strong>Fecha de Actualización:</strong> {solicitud.fecha_actualizacion.strftime('%d/%m/%Y %H:%M')}</p>
            </div>

            <div style="background-color: #e7f3ff; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h4>¿Qué significa este estado?</h4>
                <ul>
                    <li><strong>Pendiente:</strong> Su solicitud ha sido recibida y está esperando revisión</li>
                    <li><strong>En Revisión:</strong> Estamos evaluando su solicitud</li>
                    <li><strong>Aprobada:</strong> Su solicitud ha sido aprobada y será atendida pronto</li>
                    <li><strong>En Progreso:</strong> Estamos trabajando en su solicitud</li>
                    <li><strong>Completada:</strong> Su solicitud ha sido finalizada exitosamente</li>
                </ul>
            </div>

            <p>Puede hacer seguimiento de su solicitud visitando:</p>
            <p><a href="{request.host_url}solicitudes/seguimiento/{solicitud.numero_solicitud}"
                  style="background-color: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                Ver Estado de la Solicitud
            </a></p>

            <p>Si tiene alguna pregunta, no dude en contactarnos.</p>

            <p>Atentamente,<br>
            Equipo de Mantenimiento GMAO</p>
        </div>
        """

        enviar_email(solicitud.email_solicitante, asunto, contenido_html)

    except Exception as e:
        print(f"Error enviando notificación de estado: {e}")
        # No fallar la operación por error de email


@solicitudes_admin_bp.route("/<int:id>/comentario", methods=["POST"])
@login_required
def agregar_comentario(id):
    """Agregar un comentario a una solicitud"""
    if current_user.rol != "Administrador":
        return jsonify({"error": "No autorizado"}), 403

    solicitud = SolicitudServicio.query.get_or_404(id)

    data = request.get_json()
    comentario = data.get("comentario", "").strip()

    if not comentario:
        return jsonify({"error": "El comentario no puede estar vacío"}), 400

    # Agregar el comentario a las observaciones internas
    observaciones_actuales = solicitud.observaciones_internas or ""
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    nuevo_comentario = f"[{timestamp}] {current_user.nombre}: {comentario}"

    if observaciones_actuales:
        solicitud.observaciones_internas = (
            observaciones_actuales + "\n\n" + nuevo_comentario
        )
    else:
        solicitud.observaciones_internas = nuevo_comentario

    solicitud.fecha_actualizacion = datetime.now(timezone.utc)

    try:
        db.session.commit()
        return jsonify({"success": True, "message": "Comentario agregado exitosamente"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


def exportar_solicitudes_csv():
    """Genera un archivo Excel con todas las solicitudes de servicio"""
    solicitudes = SolicitudServicio.query.order_by(
        SolicitudServicio.fecha_creacion.desc()
    ).all()

    # Crear un nuevo workbook de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitudes de Servicio"

    # Estilos para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = [
        "Número Solicitud",
        "Fecha Creación",
        "Fecha Actualización",
        "Estado",
        "Prioridad",
        "Tipo Servicio",
        "Título",
        "Descripción",
        "Nombre Solicitante",
        "Email Solicitante",
        "Teléfono Solicitante",
        "Empresa Solicitante",
        "Ubicación",
        "Activo Afectado",
        "Costo Estimado",
        "Tiempo Estimado",
        "Observaciones Internas",
    ]

    # Aplicar estilos al encabezado
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ajustar ancho de columnas
    column_widths = [18, 18, 18, 12, 10, 15, 30, 40, 20, 25, 15, 20, 20, 20, 15, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Escribir datos
    for row_num, solicitud in enumerate(solicitudes, 2):
        ws.cell(row=row_num, column=1, value=solicitud.numero_solicitud)
        ws.cell(
            row=row_num,
            column=2,
            value=(
                solicitud.fecha_creacion.strftime("%d/%m/%Y %H:%M")
                if solicitud.fecha_creacion
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=3,
            value=(
                solicitud.fecha_actualizacion.strftime("%d/%m/%Y %H:%M")
                if solicitud.fecha_actualizacion
                else ""
            ),
        )
        ws.cell(row=row_num, column=4, value=solicitud.estado_display)
        ws.cell(
            row=row_num,
            column=5,
            value=solicitud.prioridad.title() if solicitud.prioridad else "",
        )
        ws.cell(
            row=row_num,
            column=6,
            value=solicitud.tipo_servicio.title() if solicitud.tipo_servicio else "",
        )
        ws.cell(row=row_num, column=7, value=solicitud.titulo)
        ws.cell(row=row_num, column=8, value=solicitud.descripcion)
        ws.cell(row=row_num, column=9, value=solicitud.nombre_solicitante)
        ws.cell(row=row_num, column=10, value=solicitud.email_solicitante)
        ws.cell(row=row_num, column=11, value=solicitud.telefono_solicitante or "")
        ws.cell(row=row_num, column=12, value=solicitud.empresa_solicitante or "")
        ws.cell(row=row_num, column=13, value=solicitud.ubicacion or "")
        ws.cell(row=row_num, column=14, value=solicitud.activo_afectado or "")
        ws.cell(
            row=row_num,
            column=15,
            value=float(solicitud.costo_estimado) if solicitud.costo_estimado else "",
        )
        ws.cell(row=row_num, column=16, value=solicitud.tiempo_estimado or "")
        ws.cell(row=row_num, column=17, value=solicitud.observaciones_internas or "")

    # Guardar el workbook en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@solicitudes_admin_bp.route("/api/filtrar", methods=["GET"])
@login_required_ajax
def filtrar_solicitudes_ajax():
    """Endpoint AJAX para filtrar solicitudes dinámicamente"""
    if current_user.rol != "Administrador":
        return jsonify({"error": "No tiene permisos para acceder a esta sección"}), 403

    # Verificar si es una petición AJAX
    if not request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"error": "Esta endpoint solo acepta peticiones AJAX"}), 400

    # Obtener filtros
    estado = request.args.get("estado", "")
    prioridad = request.args.get("prioridad", "")
    tipo_servicio = request.args.get("tipo_servicio", "")
    busqueda = request.args.get("busqueda", "").strip()
    page = int(request.args.get("page", 1))
    per_page = 25

    # Construir query
    query = SolicitudServicio.query

    if estado:
        query = query.filter_by(estado=estado)

    if prioridad:
        query = query.filter_by(prioridad=prioridad)

    if tipo_servicio:
        query = query.filter_by(tipo_servicio=tipo_servicio)

    if busqueda:
        # Buscar en número de solicitud, nombre, email o teléfono
        query = query.filter(
            db.or_(
                SolicitudServicio.numero_solicitud.ilike(f"%{busqueda}%"),
                SolicitudServicio.nombre_solicitante.ilike(f"%{busqueda}%"),
                SolicitudServicio.email_solicitante.ilike(f"%{busqueda}%"),
                SolicitudServicio.telefono_solicitante.ilike(f"%{busqueda}%"),
            )
        )

    # Ordenar por fecha de creación (más recientes primero)
    query = query.order_by(SolicitudServicio.fecha_creacion.desc())

    # Paginación
    solicitudes = query.paginate(page=page, per_page=per_page, error_out=False)

    # Convertir resultados a JSON
    solicitudes_data = []
    for solicitud in solicitudes.items:
        solicitudes_data.append(
            {
                "id": solicitud.id,
                "numero_solicitud": solicitud.numero_solicitud,
                "nombre_solicitante": solicitud.nombre_solicitante,
                "email_solicitante": solicitud.email_solicitante,
                "telefono_solicitante": solicitud.telefono_solicitante,
                "tipo_servicio": solicitud.tipo_servicio,
                "prioridad": solicitud.prioridad,
                "estado": solicitud.estado,
                "fecha_creacion": (
                    solicitud.fecha_creacion.strftime("%d/%m/%Y %H:%M")
                    if solicitud.fecha_creacion
                    else ""
                ),
                "descripcion": (
                    solicitud.descripcion[:100] + "..."
                    if solicitud.descripcion and len(solicitud.descripcion) > 100
                    else solicitud.descripcion or ""
                ),
            }
        )

    return jsonify(
        {
            "solicitudes": solicitudes_data,
            "total": solicitudes.total,
            "pages": solicitudes.pages,
            "current_page": solicitudes.page,
            "has_next": solicitudes.has_next,
            "has_prev": solicitudes.has_prev,
            "next_page": solicitudes.next_num if solicitudes.has_next else None,
            "prev_page": solicitudes.prev_num if solicitudes.has_prev else None,
        }
    )


@solicitudes_admin_bp.route("/exportar", methods=["GET"])
@login_required
def exportar_solicitudes():
    """Exporta todas las solicitudes a Excel"""
    if current_user.rol != "Administrador":
        flash("No tiene permisos para acceder a esta sección.", "error")
        return redirect(url_for("web.index"))

    try:
        excel_data = exportar_solicitudes_csv()

        response = Response(
            excel_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-disposition": "attachment; filename=solicitudes_servicio.xlsx"
            },
        )
        return response
    except Exception as e:
        flash(f"Error al exportar solicitudes: {str(e)}", "error")
        return redirect(url_for("solicitudes_admin.listar_solicitudes"))


@solicitudes_admin_bp.route("/<int:id>/eliminar", methods=["DELETE"])
@login_required_ajax
def eliminar_solicitud(id):
    """Elimina una solicitud específica"""
    if current_user.rol != "Administrador":
        return jsonify({"error": "No tiene permisos para eliminar solicitudes."}), 403

    try:
        solicitud = SolicitudServicio.query.get_or_404(id)
        
        # Eliminar archivos adjuntos asociados
        archivos = ArchivoAdjunto.query.filter_by(solicitud_id=id).all()
        for archivo in archivos:
            # Eliminar archivo físico si existe
            if archivo.ruta_archivo and os.path.exists(archivo.ruta_archivo):
                try:
                    os.remove(archivo.ruta_archivo)
                except OSError:
                    pass  # Continuar aunque no se pueda eliminar el archivo físico
            
            # Eliminar registro de la base de datos
            db.session.delete(archivo)
        
        # Eliminar la solicitud
        db.session.delete(solicitud)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"Solicitud #{solicitud.id} eliminada correctamente."
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "error": f"Error al eliminar la solicitud: {str(e)}"
        }), 500


@solicitudes_admin_bp.route("/eliminar-masivo", methods=["DELETE"])
@login_required_ajax
def eliminar_solicitudes_masivo():
    """Elimina múltiples solicitudes seleccionadas"""
    if current_user.rol != "Administrador":
        return jsonify({"error": "No tiene permisos para eliminar solicitudes."}), 403

    try:
        data = request.get_json()
        if not data or 'ids' not in data:
            return jsonify({"error": "No se proporcionaron IDs de solicitudes."}), 400
        
        ids = data['ids']
        if not isinstance(ids, list) or not ids:
            return jsonify({"error": "Lista de IDs inválida."}), 400
        
        eliminadas = 0
        errores = []
        
        for solicitud_id in ids:
            try:
                solicitud = SolicitudServicio.query.get(solicitud_id)
                if not solicitud:
                    errores.append(f"Solicitud #{solicitud_id} no encontrada")
                    continue
                
                # Eliminar archivos adjuntos asociados
                archivos = ArchivoAdjunto.query.filter_by(solicitud_id=solicitud_id).all()
                for archivo in archivos:
                    # Eliminar archivo físico si existe
                    if archivo.ruta_archivo and os.path.exists(archivo.ruta_archivo):
                        try:
                            os.remove(archivo.ruta_archivo)
                        except OSError:
                            pass  # Continuar aunque no se pueda eliminar el archivo físico
                    
                    # Eliminar registro de la base de datos
                    db.session.delete(archivo)
                
                # Eliminar la solicitud
                db.session.delete(solicitud)
                eliminadas += 1
                
            except Exception as e:
                errores.append(f"Error al eliminar solicitud #{solicitud_id}: {str(e)}")
        
        db.session.commit()
        
        mensaje = f"Se eliminaron {eliminadas} solicitudes correctamente."
        if errores:
            mensaje += f" Errores: {'; '.join(errores)}"
        
        return jsonify({
            "success": True,
            "message": mensaje,
            "eliminadas": eliminadas,
            "errores": errores
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "error": f"Error en la eliminación masiva: {str(e)}"
        }), 500
