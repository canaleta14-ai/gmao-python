from app.models.orden_trabajo import OrdenTrabajo
from app.models.activo import Activo
from app.models.usuario import Usuario
from app.models.plan_mantenimiento import PlanMantenimiento
from app.controllers.planes_controller import calcular_proxima_ejecucion
from app.extensions import db
from datetime import datetime, timezone
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def listar_ordenes(estado=None, limit=None):
    """Listar órdenes de trabajo con filtro opcional por estado y límite"""
    from sqlalchemy.orm import joinedload

    # Cargar las relaciones de forma eager para evitar N+1 queries
    query = OrdenTrabajo.query.options(
        joinedload(OrdenTrabajo.activo), joinedload(OrdenTrabajo.tecnico)
    )

    if estado:
        query = query.filter_by(estado=estado)

    # Ordenar por número de orden (ascendente) - de la OT-001 en adelante
    query = query.order_by(OrdenTrabajo.numero_orden.asc())

    if limit:
        try:
            query = query.limit(int(limit))
        except ValueError:
            pass  # Ignorar límite inválido

    ordenes = query.all()

    return [
        {
            "id": o.id,
            "numero_orden": o.numero_orden,
            "fecha_creacion": (
                o.fecha_creacion.strftime("%d/%m/%Y %H:%M")
                if o.fecha_creacion
                else None
            ),
            "fecha_programada": (
                o.fecha_programada.strftime("%d/%m/%Y") if o.fecha_programada else None
            ),
            "fecha_completada": (
                o.fecha_completada.strftime("%d/%m/%Y %H:%M")
                if o.fecha_completada
                else None
            ),
            "tipo": o.tipo,
            "prioridad": o.prioridad,
            "estado": o.estado,
            "descripcion": o.descripcion,
            "observaciones": o.observaciones,
            "tiempo_estimado": o.tiempo_estimado,
            "tiempo_real": o.tiempo_real,
            "activo_id": o.activo_id,
            "activo_nombre": o.activo.nombre if o.activo else None,
            "activo_codigo": o.activo.codigo if o.activo else None,
            "tecnico_id": o.tecnico_id,
            "tecnico_nombre": o.tecnico.nombre if o.tecnico else None,
        }
        for o in ordenes
    ]


def listar_ordenes_paginado(
    page=1, per_page=10, q=None, estado=None, tipo=None, prioridad=None
):
    """Listar órdenes de trabajo con paginación y filtros"""
    from sqlalchemy.orm import joinedload

    # Cargar las relaciones de forma eager para evitar N+1 queries
    query = OrdenTrabajo.query.options(
        joinedload(OrdenTrabajo.activo), joinedload(OrdenTrabajo.tecnico)
    )

    # Filtro por estado
    if estado:
        query = query.filter_by(estado=estado)

    # Filtro por tipo
    if tipo:
        query = query.filter_by(tipo=tipo)

    # Filtro por prioridad
    if prioridad:
        query = query.filter_by(prioridad=prioridad)

    # Filtro de búsqueda general
    if q:
        search_term = f"%{q}%"
        query = query.join(Activo, OrdenTrabajo.activo_id == Activo.id, isouter=True)
        query = query.join(Usuario, OrdenTrabajo.tecnico_id == Usuario.id, isouter=True)
        query = query.filter(
            db.or_(
                OrdenTrabajo.numero_orden.ilike(search_term),
                OrdenTrabajo.descripcion.ilike(search_term),
                OrdenTrabajo.tipo.ilike(search_term),
                OrdenTrabajo.estado.ilike(search_term),
                OrdenTrabajo.prioridad.ilike(search_term),
                Activo.nombre.ilike(search_term),
                Activo.codigo.ilike(search_term),
                Usuario.nombre.ilike(search_term),
            )
        )

    # Ordenamiento por número de orden ascendente (OT-001, OT-002, etc.)
    query = query.order_by(OrdenTrabajo.numero_orden.asc())

    # Paginación
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    ordenes_data = [
        {
            "id": o.id,
            "numero_orden": o.numero_orden,
            "fecha_creacion": (
                o.fecha_creacion.strftime("%d/%m/%Y %H:%M")
                if o.fecha_creacion
                else None
            ),
            "fecha_programada": (
                o.fecha_programada.strftime("%d/%m/%Y") if o.fecha_programada else None
            ),
            "fecha_completada": (
                o.fecha_completada.strftime("%d/%m/%Y %H:%M")
                if o.fecha_completada
                else None
            ),
            "tipo": o.tipo,
            "prioridad": o.prioridad,
            "estado": o.estado,
            "descripcion": o.descripcion,
            "observaciones": o.observaciones,
            "tiempo_estimado": o.tiempo_estimado,
            "tiempo_real": o.tiempo_real,
            "activo_id": o.activo_id,
            "activo_nombre": o.activo.nombre if o.activo else None,
            "activo_codigo": o.activo.codigo if o.activo else None,
            "tecnico_id": o.tecnico_id,
            "tecnico_nombre": o.tecnico.nombre if o.tecnico else None,
        }
        for o in pagination.items
    ]

    return {
        "items": ordenes_data,
        "page": pagination.page,
        "per_page": pagination.per_page,
        "total": pagination.total,
        "pages": pagination.pages,
        "has_next": pagination.has_next,
        "has_prev": pagination.has_prev,
    }


def obtener_orden(id):
    """Obtener una orden de trabajo específica"""
    orden = OrdenTrabajo.query.get_or_404(id)
    return {
        "id": orden.id,
        "numero_orden": orden.numero_orden,
        "fecha_creacion": (
            orden.fecha_creacion.strftime("%Y-%m-%dT%H:%M")
            if orden.fecha_creacion
            else None
        ),
        "fecha_programada": (
            orden.fecha_programada.strftime("%Y-%m-%d")
            if orden.fecha_programada
            else None
        ),
        "fecha_completada": (
            orden.fecha_completada.strftime("%Y-%m-%dT%H:%M")
            if orden.fecha_completada
            else None
        ),
        "tipo": orden.tipo,
        "prioridad": orden.prioridad,
        "estado": orden.estado,
        "descripcion": orden.descripcion,
        "observaciones": orden.observaciones,
        "tiempo_estimado": orden.tiempo_estimado,
        "tiempo_real": orden.tiempo_real,
        "activo_id": orden.activo_id,
        "tecnico_id": orden.tecnico_id,
    }


def crear_orden(data):
    """Crear nueva orden de trabajo"""
    # Generar número de orden único
    ultimo_numero = db.session.query(db.func.max(OrdenTrabajo.id)).scalar() or 0
    numero_orden = f"OT-{ultimo_numero + 1:06d}"

    # Validar que el activo existe si se proporciona
    if data.get("activo_id"):
        activo = Activo.query.get(data["activo_id"])
        if not activo:
            raise ValueError("El activo especificado no existe")

    # Validar que el técnico existe si se proporciona
    if data.get("tecnico_id"):
        tecnico = Usuario.query.get(data["tecnico_id"])
        if not tecnico:
            raise ValueError("El técnico especificado no existe")

    nueva_orden = OrdenTrabajo(
        numero_orden=numero_orden,
        tipo=data["tipo"],
        prioridad=data["prioridad"],
        estado="Pendiente",
        descripcion=data["descripcion"],
        observaciones=data.get("observaciones"),
        activo_id=data.get("activo_id"),
        tecnico_id=data.get("tecnico_id"),
        tiempo_estimado=(
            float(data.get("tiempo_estimado", 0))
            if data.get("tiempo_estimado")
            else None
        ),
    )

    # Procesar fecha programada
    if data.get("fecha_programada"):
        try:
            nueva_orden.fecha_programada = datetime.strptime(
                data["fecha_programada"], "%Y-%m-%d"
            )
        except ValueError:
            raise ValueError("Formato de fecha programada inválido. Use YYYY-MM-DD")

    db.session.add(nueva_orden)
    db.session.commit()
    return nueva_orden


def actualizar_orden(id, data):
    """Actualizar orden de trabajo existente"""
    orden = OrdenTrabajo.query.get_or_404(id)

    # Validar que el activo existe si se proporciona
    if data.get("activo_id") is not None and data["activo_id"] != orden.activo_id:
        if data["activo_id"]:  # Solo validar si no es None o 0
            activo = Activo.query.get(data["activo_id"])
            if not activo:
                raise ValueError(f"El activo con ID {data['activo_id']} no existe")

    # Validar que el técnico existe si se proporciona
    if data.get("tecnico_id") is not None and data["tecnico_id"] != orden.tecnico_id:
        if data["tecnico_id"]:  # Solo validar si no es None o 0
            tecnico = Usuario.query.get(data["tecnico_id"])
            if not tecnico:
                raise ValueError(f"El técnico con ID {data['tecnico_id']} no existe")

    # Validar campos requeridos
    if not data.get("tipo"):
        raise ValueError("El tipo de orden es requerido")

    if not data.get("prioridad"):
        raise ValueError("La prioridad es requerida")

    if not data.get("descripcion") or data.get("descripcion").strip() == "":
        raise ValueError("La descripción es requerida")

    # Actualizar campos
    orden.tipo = data.get("tipo", orden.tipo)
    orden.prioridad = data.get("prioridad", orden.prioridad)
    orden.descripcion = data.get("descripcion", orden.descripcion)
    orden.observaciones = data.get("observaciones", orden.observaciones)
    orden.activo_id = data.get("activo_id", orden.activo_id)
    orden.tecnico_id = data.get("tecnico_id", orden.tecnico_id)

    # Actualizar tiempo estimado
    if data.get("tiempo_estimado") is not None:
        try:
            orden.tiempo_estimado = (
                float(data["tiempo_estimado"]) if data["tiempo_estimado"] else None
            )
        except (ValueError, TypeError):
            raise ValueError("El tiempo estimado debe ser un número válido")

    # Procesar fecha programada
    if data.get("fecha_programada") is not None:
        if data["fecha_programada"]:
            try:
                orden.fecha_programada = datetime.strptime(
                    data["fecha_programada"], "%Y-%m-%d"
                )
            except ValueError:
                raise ValueError("Formato de fecha programada inválido. Use YYYY-MM-DD")
        else:
            orden.fecha_programada = None

    try:
        db.session.commit()
        return orden
    except Exception as e:
        db.session.rollback()
        raise ValueError(f"Error al guardar en la base de datos: {str(e)}")


def actualizar_estado_orden(id, data):
    """Actualizar estado de una orden de trabajo"""
    orden = OrdenTrabajo.query.get_or_404(id)

    estado_anterior = orden.estado
    nuevo_estado = data["estado"]

    # Validar transición de estado
    estados_validos = [
        "Pendiente",
        "En Proceso",
        "Completada",
        "Cancelada",
        "En Espera",
    ]
    if nuevo_estado not in estados_validos:
        raise ValueError(f"Estado '{nuevo_estado}' no válido")

    orden.estado = nuevo_estado

    # Lógica especial para estados específicos
    if nuevo_estado == "Completada" and estado_anterior != "Completada":
        orden.fecha_completada = datetime.now(timezone.utc)
        # Actualizar tiempo real si se proporciona
        if data.get("tiempo_real"):
            orden.tiempo_real = float(data["tiempo_real"])

        # Si es una orden preventiva, actualizar la próxima ejecución del plan correspondiente
        if orden.tipo and "preventivo" in orden.tipo.lower() and orden.activo_id:
            try:
                # Buscar planes activos para este activo
                planes_activos = PlanMantenimiento.query.filter_by(
                    activo_id=orden.activo_id, estado="Activo"
                ).all()

                if planes_activos:
                    # Encontrar el plan más próximo a vencer (o ya vencido)
                    fecha_actual = datetime.now()
                    plan_a_actualizar = None
                    diferencia_minima = None

                    for plan in planes_activos:
                        if plan.proxima_ejecucion:
                            diferencia = abs(
                                (plan.proxima_ejecucion - fecha_actual).total_seconds()
                            )
                            if (
                                diferencia_minima is None
                                or diferencia < diferencia_minima
                            ):
                                diferencia_minima = diferencia
                                plan_a_actualizar = plan

                    # Si encontramos un plan, actualizar su próxima ejecución
                    if plan_a_actualizar:
                        # Preparar los datos de configuración del plan para recalcular
                        plan_data = {
                            "tipo_frecuencia": plan_a_actualizar.tipo_frecuencia
                            or "mensual",
                            "intervalo_meses": plan_a_actualizar.intervalo_meses or 1,
                            "tipo_mensual": plan_a_actualizar.tipo_mensual or "dia_mes",
                            "dia_mes": plan_a_actualizar.dia_mes,
                            "semana_mes": plan_a_actualizar.semana_mes,
                            "dia_semana_mes": plan_a_actualizar.dia_semana_mes,
                            "intervalo_semanas": plan_a_actualizar.intervalo_semanas
                            or 1,
                            "dias_semana": plan_a_actualizar.dias_semana,
                            "frecuencia": plan_a_actualizar.frecuencia,
                        }

                        # Calcular nueva próxima ejecución usando la fecha de completado como base
                        nueva_proxima_ejecucion = calcular_proxima_ejecucion(
                            plan_data, orden.fecha_completada
                        )

                        # Actualizar el plan
                        plan_a_actualizar.ultima_ejecucion = orden.fecha_completada
                        plan_a_actualizar.proxima_ejecucion = nueva_proxima_ejecucion

                        print(
                            f"DEBUG: Actualizado plan {plan_a_actualizar.nombre} - Nueva próxima ejecución: {nueva_proxima_ejecucion}"
                        )

            except Exception as e:
                print(f"Error al actualizar plan de mantenimiento: {e}")
                # No fallar la actualización de la orden por errores en planes

    elif nuevo_estado != "Completada":
        # Si se cambia de Completada a otro estado, limpiar fecha de completado
        orden.fecha_completada = None

    # Actualizar observaciones si se proporcionan
    if data.get("observaciones"):
        orden.observaciones = data["observaciones"]

    db.session.commit()
    return orden


def eliminar_orden(id):
    """Eliminar orden de trabajo"""
    orden = OrdenTrabajo.query.get_or_404(id)

    # Permitir eliminar órdenes en estado Pendiente, En Proceso o Cancelada
    if orden.estado not in ["Pendiente", "En Proceso", "Cancelada"]:
        raise ValueError(
            "Solo se pueden eliminar órdenes en estado 'Pendiente', 'En Proceso' o 'Cancelada'"
        )

    db.session.delete(orden)
    db.session.commit()
    return True


def obtener_activos_disponibles():
    """Obtener lista de activos para selección en órdenes"""
    activos = Activo.query.filter_by(estado="Operativo").order_by(Activo.codigo).all()
    return [
        {
            "id": activo.id,
            "codigo": activo.codigo,
            "nombre": activo.nombre,
            "ubicacion": activo.ubicacion,
            "display": f"{activo.codigo} - {activo.nombre}"
            + (f" ({activo.ubicacion})" if activo.ubicacion else ""),
        }
        for activo in activos
    ]


def obtener_tecnicos_disponibles():
    """Obtener lista de técnicos para asignación de órdenes"""
    tecnicos = (
        Usuario.query.filter_by(activo=True)
        .filter(Usuario.rol.in_(["Técnico", "Supervisor", "Administrador"]))
        .order_by(Usuario.nombre)
        .all()
    )

    return [
        {
            "id": tecnico.id,
            "nombre": tecnico.nombre,
            "rol": tecnico.rol,
            "display": f"{tecnico.nombre} ({tecnico.rol})",
        }
        for tecnico in tecnicos
    ]


def obtener_estadisticas_ordenes():
    """Obtener estadísticas de órdenes de trabajo"""
    from sqlalchemy import func, and_
    from datetime import date, timedelta

    # Contar por estado
    stats_estado = (
        db.session.query(OrdenTrabajo.estado, func.count(OrdenTrabajo.id))
        .group_by(OrdenTrabajo.estado)
        .all()
    )

    # Órdenes de esta semana
    inicio_semana = date.today() - timedelta(days=7)
    ordenes_semana = OrdenTrabajo.query.filter(
        OrdenTrabajo.fecha_creacion >= inicio_semana
    ).count()

    # Órdenes vencidas (programadas para antes de hoy y no completadas)
    hoy = date.today()
    ordenes_vencidas = OrdenTrabajo.query.filter(
        and_(OrdenTrabajo.fecha_programada < hoy, OrdenTrabajo.estado != "Completada")
    ).count()

    # Promedio de tiempo de resolución (últimas 30 órdenes completadas)
    ordenes_completadas = (
        OrdenTrabajo.query.filter_by(estado="Completada")
        .filter(OrdenTrabajo.tiempo_real.isnot(None))
        .order_by(OrdenTrabajo.fecha_completada.desc())
        .limit(30)
        .all()
    )

    tiempo_promedio = 0
    if ordenes_completadas:
        tiempo_total = sum(o.tiempo_real for o in ordenes_completadas)
        tiempo_promedio = tiempo_total / len(ordenes_completadas)

    return {
        "por_estado": dict(stats_estado),
        "ordenes_semana": ordenes_semana,
        "ordenes_vencidas": ordenes_vencidas,
        "tiempo_promedio_resolucion": round(tiempo_promedio, 2),
    }


def exportar_ordenes_csv():
    """Exportar órdenes de trabajo a Excel"""
    ordenes = OrdenTrabajo.query.order_by(OrdenTrabajo.fecha_creacion.desc()).all()

    # Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Órdenes de Trabajo"

    # Estilos para el encabezado
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = [
        "Número",
        "Fecha Creación",
        "Fecha Programada",
        "Fecha Completada",
        "Tipo",
        "Prioridad",
        "Estado",
        "Descripción",
        "Observaciones",
        "Tiempo Estimado",
        "Tiempo Real",
        "Activo",
        "Técnico",
    ]

    # Aplicar encabezados con estilo
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Datos
    for row_num, orden in enumerate(ordenes, 2):
        ws.cell(row=row_num, column=1, value=orden.numero_orden)
        ws.cell(
            row=row_num,
            column=2,
            value=(
                orden.fecha_creacion.strftime("%d/%m/%Y %H:%M")
                if orden.fecha_creacion
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=3,
            value=(
                orden.fecha_programada.strftime("%d/%m/%Y")
                if orden.fecha_programada
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=4,
            value=(
                orden.fecha_completada.strftime("%d/%m/%Y %H:%M")
                if orden.fecha_completada
                else ""
            ),
        )
        ws.cell(row=row_num, column=5, value=orden.tipo or "")
        ws.cell(row=row_num, column=6, value=orden.prioridad or "")
        ws.cell(row=row_num, column=7, value=orden.estado or "")
        ws.cell(row=row_num, column=8, value=orden.descripcion or "")
        ws.cell(row=row_num, column=9, value=orden.observaciones or "")
        ws.cell(row=row_num, column=10, value=orden.tiempo_estimado or "")
        ws.cell(row=row_num, column=11, value=orden.tiempo_real or "")
        ws.cell(
            row=row_num, column=12, value=orden.activo.nombre if orden.activo else ""
        )
        ws.cell(
            row=row_num, column=13, value=orden.tecnico.nombre if orden.tecnico else ""
        )

    # Ajustar ancho de columnas
    column_widths = [15, 20, 18, 20, 12, 12, 12, 40, 40, 18, 15, 25, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
