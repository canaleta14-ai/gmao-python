from flask import Blueprint, request, jsonify, render_template, Response
from flask_login import login_required
from app.extensions import csrf
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

usuarios_bp = Blueprint("usuarios", __name__, url_prefix="/usuarios")


@usuarios_bp.route("/")
@login_required
def usuarios_page():
    try:
        return render_template("usuarios/usuarios.html", section="usuarios")
    except Exception:
        # Fallback seguro para evitar 500 si falta template o hay error en dependencias
        html = """
        <!DOCTYPE html>
        <html lang=\"es\">
        <head><meta charset=\"utf-8\"><title>Usuarios</title></head>
        <body>
            <h1>Usuarios</h1>
            <p>La página de usuarios no pudo renderizarse completamente, pero el sistema está operativo.</p>
        </body>
        </html>
        """
        return Response(html, mimetype="text/html")


@usuarios_bp.route("/api", methods=["GET"])
@login_required
def api_usuarios():
    try:
        from app.models.usuario import Usuario
        from app.extensions import db

        # Obtener parámetros de filtrado
        q = request.args.get("q", "").strip()
        rol = request.args.get("rol", "").strip()
        cargo = request.args.get("cargo", "").strip()
        estado = request.args.get("estado", "").strip()

        # Construir query base
        query = Usuario.query

        # Aplicar filtros
        if q:
            search_term = f"%{q}%"
            query = query.filter(
                db.or_(
                    Usuario.nombre.ilike(search_term),
                    Usuario.email.ilike(search_term),
                    Usuario.username.ilike(search_term),
                )
            )

        if rol:
            query = query.filter_by(rol=rol)

        if estado:
            activo = estado == "Activo"
            query = query.filter_by(activo=activo)

        # Obtener usuarios filtrados
        usuarios_db = query.all()

        usuarios_data = []
        for usuario in usuarios_db:
            cargo_usuario = (
                "Técnico" if usuario.rol != "Administrador" else "Administrador"
            )

            # Aplicar filtro de cargo si existe
            if cargo and cargo.lower() not in cargo_usuario.lower():
                continue

            usuarios_data.append(
                {
                    "id": usuario.id,
                    "codigo": f"USR{str(usuario.id).zfill(3)}",
                    "nombre": usuario.nombre,
                    "email": usuario.email,
                    "telefono": "",
                    "departamento": "Mantenimiento",
                    "cargo": cargo_usuario,
                    "estado": "Activo" if usuario.activo else "Inactivo",
                    "fecha_ingreso": (
                        usuario.fecha_creacion.strftime("%Y-%m-%d")
                        if usuario.fecha_creacion
                        else "2023-01-01"
                    ),
                    "rol": usuario.rol,
                    "permisos": (
                        ["*"]
                        if usuario.rol == "Administrador"
                        else ["mantenimiento.leer"]
                    ),
                }
            )

        # Compatibilidad de formato según query param
        fmt = (request.args.get("format") or "").strip().lower()
        # Alternativa: devolver lista si se pide explícitamente
        if fmt in ["list", "array"]:
            return jsonify(usuarios_data)

        # Formato por defecto: objeto con metadata
        return jsonify(
            {"success": True, "usuarios": usuarios_data, "total": len(usuarios_data)}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@usuarios_bp.route("/api", methods=["POST"])
@login_required
def crear_usuario_api():
    try:
        from app.models.usuario import Usuario
        from app.extensions import db

        print(f"🔍 DEBUG: Request content type: {request.content_type}")
        print(f"🔍 DEBUG: Request data: {request.data}")
        print(f"🔍 DEBUG: Request headers: {dict(request.headers)}")

        # Check if request has JSON data
        if not request.is_json:
            error_msg = "Request must be JSON"
            print(f"❌ DEBUG: Content type error: {error_msg}")
            return jsonify({"success": False, "error": error_msg}), 400

        data = request.get_json()
        print(f"🔍 DEBUG: Parsed JSON data: {data}")

        # Check if data is None or empty
        if not data:
            error_msg = "No JSON data provided"
            print(f"❌ DEBUG: No data error: {error_msg}")
            return jsonify({"success": False, "error": error_msg}), 400

        # Validar campos requeridos
        required_fields = ["username", "email", "password", "nombre"]
        for field in required_fields:
            if not data.get(field):
                error_msg = f"El campo {field} es requerido"
                print(f"❌ DEBUG: Validation error: {error_msg}")
                return (
                    jsonify({"success": False, "error": error_msg}),
                    400,
                )

        # Verificar que el username no exista
        if Usuario.query.filter_by(username=data["username"]).first():
            return (
                jsonify({"success": False, "error": "El nombre de usuario ya existe"}),
                400,
            )

        # Verificar que el email no exista
        if Usuario.query.filter_by(email=data["email"]).first():
            return (
                jsonify({"success": False, "error": "El email ya está registrado"}),
                400,
            )

        # Crear el nuevo usuario
        nuevo_usuario = Usuario(
            username=data["username"],
            email=data["email"],
            nombre=data["nombre"],
            rol=data.get("rol", "Técnico"),
            activo=data.get("activo", True),
        )

        # Establecer la contraseña encriptada
        nuevo_usuario.set_password(data["password"])

        # Guardar en la base de datos
        db.session.add(nuevo_usuario)
        db.session.commit()

        return jsonify(
            {
                "success": True,
                "message": "Usuario creado correctamente",
                "id": nuevo_usuario.id,
            }
        )
    except Exception as e:
        print(f"❌ DEBUG: Exception occurred: {type(e).__name__}: {str(e)}")
        db.session.rollback()

        # Check if it's a CSRF error
        if "CSRF" in str(e) or "csrf" in str(e).lower():
            return (
                jsonify({"success": False, "error": "CSRF token validation failed"}),
                400,
            )

        return jsonify({"success": False, "error": str(e)}), 500


@usuarios_bp.route("/api/<int:user_id>", methods=["GET"])
@login_required
def obtener_usuario_api(user_id):
    try:
        from app.models.usuario import Usuario

        # Buscar el usuario
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return jsonify({"success": False, "error": "Usuario no encontrado"}), 404

        # Retornar datos del usuario
        usuario_data = {
            "id": usuario.id,
            "codigo": f"USR{str(usuario.id).zfill(3)}",
            "username": usuario.username,
            "nombre": usuario.nombre,
            "email": usuario.email,
            "telefono": "",  # No tenemos teléfono en la DB actual
            "departamento": "Mantenimiento",  # Valor por defecto
            "cargo": ("Técnico" if usuario.rol != "Administrador" else "Administrador"),
            "estado": "Activo" if usuario.activo else "Inactivo",
            "fecha_ingreso": (
                usuario.fecha_creacion.strftime("%Y-%m-%d")
                if usuario.fecha_creacion
                else "2023-01-01"
            ),
            "rol": usuario.rol,
            "permisos": (
                ["*"] if usuario.rol == "Administrador" else ["mantenimiento.leer"]
            ),
        }

        return jsonify({"success": True, "usuario": usuario_data})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@usuarios_bp.route("/api/<int:user_id>", methods=["PUT"])
@login_required
def actualizar_usuario_api(user_id):
    try:
        from app.models.usuario import Usuario
        from app.extensions import db

        data = request.get_json()

        # Buscar el usuario
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return jsonify({"success": False, "error": "Usuario no encontrado"}), 404

        # Validar campos requeridos
        required_fields = ["username", "email", "nombre"]
        for field in required_fields:
            if not data.get(field):
                return (
                    jsonify(
                        {"success": False, "error": f"El campo {field} es requerido"}
                    ),
                    400,
                )

        # Verificar que el username no exista (excepto para el usuario actual)
        existing_user = Usuario.query.filter_by(username=data["username"]).first()
        if existing_user and existing_user.id != user_id:
            return (
                jsonify({"success": False, "error": "El nombre de usuario ya existe"}),
                400,
            )

        # Verificar que el email no exista (excepto para el usuario actual)
        existing_user = Usuario.query.filter_by(email=data["email"]).first()
        if existing_user and existing_user.id != user_id:
            return (
                jsonify({"success": False, "error": "El email ya está registrado"}),
                400,
            )

        # Actualizar los campos
        usuario.username = data["username"]
        usuario.email = data["email"]
        usuario.nombre = data["nombre"]
        usuario.rol = data.get("rol", "Técnico")

        # Convertir estado "Activo"/"Inactivo" a booleano
        estado = data.get("estado", "Activo")
        usuario.activo = estado == "Activo"

        # Actualizar contraseña solo si se proporcionó
        if data.get("password"):
            usuario.set_password(data["password"])

        # Guardar cambios
        db.session.commit()

        return jsonify(
            {
                "success": True,
                "message": f"Usuario {usuario.nombre} actualizado correctamente",
            }
        )
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@usuarios_bp.route("/api/<int:user_id>", methods=["DELETE"])
@login_required
def eliminar_usuario_api(user_id):
    try:
        from app.models.usuario import Usuario
        from app.extensions import db

        # Buscar el usuario en la base de datos
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return jsonify({"success": False, "error": "Usuario no encontrado"}), 404

        # No permitir eliminar el usuario admin
        if usuario.username == "admin":
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "No se puede eliminar el usuario administrador",
                    }
                ),
                403,
            )

        # Eliminar el usuario
        nombre_usuario = usuario.nombre
        db.session.delete(usuario)
        db.session.commit()

        return jsonify(
            {
                "success": True,
                "message": f"Usuario {nombre_usuario} eliminado correctamente",
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@usuarios_bp.route("/api/roles", methods=["GET"])
@login_required
def api_roles():
    try:
        roles_data = [
            {
                "id": 1,
                "nombre": "Administrador",
                "descripcion": "Acceso completo al sistema",
                "permisos": ["*"],
                "color": "danger",
            },
            {
                "id": 2,
                "nombre": "Supervisor",
                "descripcion": "Supervisión y gestión de equipos",
                "permisos": [
                    "usuarios.leer",
                    "activos.leer",
                    "ordenes.leer",
                    "operaciones.leer",
                    "operaciones.editar",
                ],
                "color": "warning",
            },
            {
                "id": 3,
                "nombre": "Técnico",
                "descripcion": "Operaciones técnicas y mantenimiento",
                "permisos": [
                    "mantenimiento.leer",
                    "mantenimiento.editar",
                    "ordenes.crear",
                    "ordenes.editar",
                    "activos.leer",
                ],
                "color": "primary",
            },
            {
                "id": 4,
                "nombre": "Operador",
                "descripcion": "Operaciones básicas",
                "permisos": ["ordenes.leer", "activos.leer"],
                "color": "secondary",
            },
        ]
        return jsonify({"success": True, "roles": roles_data, "total": len(roles_data)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@usuarios_bp.route("/exportar-csv", methods=["GET"])
@login_required
def exportar_csv():
    """Exporta todos los usuarios a Excel"""
    try:
        from app.models.usuario import Usuario

        # Obtener todos los usuarios
        usuarios = Usuario.query.all()

        # Crear workbook y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        # Estilos para el encabezado
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = [
            "ID",
            "Código",
            "Nombre",
            "Email",
            "Teléfono",
            "Departamento",
            "Cargo",
            "Rol",
            "Estado",
            "Fecha Ingreso",
        ]

        # Aplicar encabezados con estilo
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Datos
        for row_num, usuario in enumerate(usuarios, 2):
            ws.cell(row=row_num, column=1, value=usuario.id)
            ws.cell(row=row_num, column=2, value=f"USR{str(usuario.id).zfill(3)}")
            ws.cell(row=row_num, column=3, value=usuario.nombre)
            ws.cell(row=row_num, column=4, value=usuario.email)
            ws.cell(row=row_num, column=5, value="")
            ws.cell(row=row_num, column=6, value="Mantenimiento")
            ws.cell(
                row=row_num,
                column=7,
                value=(
                    "Técnico" if usuario.rol != "Administrador" else "Administrador"
                ),
            )
            ws.cell(row=row_num, column=8, value=usuario.rol)
            ws.cell(
                row=row_num, column=9, value="Activo" if usuario.activo else "Inactivo"
            )
            ws.cell(
                row=row_num,
                column=10,
                value=(
                    usuario.fecha_creacion.strftime("%Y-%m-%d")
                    if usuario.fecha_creacion
                    else "2023-01-01"
                ),
            )

        # Ajustar ancho de columnas
        column_widths = [8, 12, 25, 30, 15, 15, 15, 12, 10, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = (
                width
            )

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = output.getvalue()
        output.close()

        response = Response(
            excel_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": "attachment; filename=usuarios.xlsx"},
        )
        return response
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Nuevos endpoints para validación en tiempo real
@usuarios_bp.route("/api/validar-username", methods=["POST"])
@login_required
@csrf.exempt  # Para AJAX
def validar_username():
    """Valida si un username está disponible"""
    try:
        from app.models.usuario import Usuario

        data = request.get_json()
        if not data or "username" not in data:
            return jsonify({"error": "Username requerido"}), 400

        username = data["username"].strip()
        usuario_id = data.get("usuario_id")  # Para edición

        if not username:
            return (
                jsonify(
                    {"disponible": False, "mensaje": "Username no puede estar vacío"}
                ),
                200,
            )

        # Buscar usuario existente
        query = Usuario.query.filter_by(username=username)
        if usuario_id:  # En edición, excluir el usuario actual
            query = query.filter(Usuario.id != usuario_id)

        usuario_existente = query.first()

        if usuario_existente:
            return (
                jsonify(
                    {
                        "disponible": False,
                        "mensaje": f"El username '{username}' ya está registrado",
                    }
                ),
                200,
            )
        else:
            return (
                jsonify(
                    {"disponible": True, "mensaje": f"Username '{username}' disponible"}
                ),
                200,
            )

    except Exception as e:
        return jsonify({"error": f"Error al validar username: {str(e)}"}), 500


@usuarios_bp.route("/api/validar-email", methods=["POST"])
@login_required
@csrf.exempt  # Para AJAX
def validar_email():
    """Valida si un email está disponible"""
    try:
        from app.models.usuario import Usuario
        import re

        data = request.get_json()
        if not data or "email" not in data:
            return jsonify({"error": "Email requerido"}), 400

        email = data["email"].strip().lower()
        usuario_id = data.get("usuario_id")  # Para edición

        if not email:
            return (
                jsonify({"disponible": False, "mensaje": "Email no puede estar vacío"}),
                200,
            )

        # Validar formato de email
        email_regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
        if not re.match(email_regex, email):
            return (
                jsonify({"disponible": False, "mensaje": "Formato de email inválido"}),
                200,
            )

        # Buscar email existente
        query = Usuario.query.filter_by(email=email)
        if usuario_id:  # En edición, excluir el usuario actual
            query = query.filter(Usuario.id != usuario_id)

        usuario_existente = query.first()

        if usuario_existente:
            return (
                jsonify(
                    {
                        "disponible": False,
                        "mensaje": f"El email '{email}' ya está registrado",
                    }
                ),
                200,
            )
        else:
            return (
                jsonify({"disponible": True, "mensaje": f"Email '{email}' disponible"}),
                200,
            )

    except Exception as e:
        return jsonify({"error": f"Error al validar email: {str(e)}"}), 500


@usuarios_bp.route("/api/autocomplete", methods=["GET"])
@login_required
def autocomplete_usuarios():
    """API para autocompletado de usuarios"""
    try:
        from app.models.usuario import Usuario
        from app.extensions import db

        # Obtener parámetro de búsqueda
        q = request.args.get("q", "").strip()
        
        # Si no hay parámetro de búsqueda o es muy corto, devolver usuarios activos limitados
        if not q or len(q) < 2:
            usuarios = Usuario.query.filter_by(activo=True).limit(10).all()
        else:
            # Buscar usuarios activos que coincidan con la búsqueda
            search_term = f"%{q}%"
            usuarios = Usuario.query.filter(
                db.and_(
                    Usuario.activo == True,
                    db.or_(
                        Usuario.username.ilike(search_term),
                        Usuario.nombre.ilike(search_term),
                        Usuario.email.ilike(search_term)
                    )
                )
            ).limit(10).all()

        # Formatear resultados para autocompletado
        resultados = []
        for usuario in usuarios:
            resultados.append({
                "id": usuario.id,
                "username": usuario.username,
                "nombre": usuario.nombre,
                "email": usuario.email,
                "rol": usuario.rol,
                "value": usuario.username,  # Valor que se mostrará en el input
                "label": f"{usuario.username} - {usuario.nombre}"  # Etiqueta descriptiva
            })

        return jsonify(resultados)

    except Exception as e:
        print(f"Error en autocomplete_usuarios: {str(e)}")
        return jsonify({"error": str(e)}), 500
