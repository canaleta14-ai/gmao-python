import os
import sqlite3
import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), "instance/database.db")
EXCEL_PATH = "Proveedores.xlsx"

# Crear tabla proveedor adaptada al Excel
conn = sqlite3.connect(DB_PATH)
c = conn.cursor()
c.execute("DROP TABLE IF EXISTS proveedor")
c.execute("""
CREATE TABLE IF NOT EXISTS proveedor (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    proveedor TEXT,
    nif TEXT,
    direccion TEXT,
    contacto TEXT,
    email TEXT,
    cuenta_contable TEXT,
    estado TEXT
)
""")

# Leer datos del Excel
wb = openpyxl.load_workbook(EXCEL_PATH)
sheet = wb.active

columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
for row in sheet.iter_rows(min_row=2, values_only=True):
    c.execute("""
        INSERT INTO proveedor (proveedor, nif, direccion, contacto, email, cuenta_contable, estado)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, row)

conn.commit()
conn.close()
print("Proveedores importados correctamente desde el Excel.")